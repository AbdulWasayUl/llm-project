import openpyxl
import json
import re

def clean_answer_list(answer_list):
    cleaned = []
    skip_count = 0
    i = 0
    while i < len(answer_list):
        item = answer_list[i]
        if (
            isinstance(item, str) and item.strip() == "Profit Payment"
            and i + 1 < len(answer_list)
            and isinstance(answer_list[i + 1], str)
            and answer_list[i + 1].strip() == "Profit Rate"
        ):
            # Skip next 5 entries total: 2 headers + 1 buffer + 2 values
            i += 5
            continue
        if item is not None:
            cleaned.append(item)
        i += 1
    return cleaned

QUESTION_WORD_REGEX = re.compile(
    r"(?i)^\s*(is|please|what|how|when|where|why|are|can|do|does|shall|should|could|would|will|who|which)\b.*\.$"
)

def is_question(text):
    if not isinstance(text, str):
        return False
    stripped = text.strip()
    if stripped.endswith("?"):
        return True
    return bool(QUESTION_WORD_REGEX.match(stripped))

def extract_qa_from_sheet(sheet):
    qas = []
    max_row = sheet.max_row
    max_col = sheet.max_column
    title = sheet.cell(row=1, column=1).value

    current_question = None
    current_answer = []

    row_idx = 2  # Start after title
    while row_idx <= max_row:
        row = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
        found_question = False

        for col_idx, cell_value in enumerate(row):
            if is_question(cell_value):
                found_question = True

                # Save the previous Q&A
                if current_question and current_answer:
                    qas.append({
                        "question": current_question.strip(),
                        "answer": [a for a in current_answer if a is not None]
                    })

                current_question = cell_value
                current_answer = []

                # Add answer to the right of the question in the same row
                right_cells = row[col_idx + 1:]
                current_answer.extend([val for val in right_cells if val is not None])
                break  # Only one question per row

        # If no question found, collect answer content
        if not found_question and current_question:
            current_answer.extend([val for val in row if val is not None])

        row_idx += 1

    # Append the last question-answer pair
    if current_question and current_answer:
        qas.append({
            "question": current_question.strip(),
            "answer": clean_answer_list(current_answer)
        })

    return {
        "title": title,
        "qas": qas
    }

def process_excel_to_json(filepath, skip=False):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheetnames = wb.sheetnames

    final_output = {}

    skipped_sheets = 0 if not skip else 2

    for sheet_name in sheetnames[skipped_sheets:]:  # Skip first two sheets
        sheet = wb[sheet_name]
        final_output[sheet_name] = extract_qa_from_sheet(sheet)

    with open("C:/Users/HP/Desktop/account_qass.json", "w", encoding="utf-8") as f:
        json.dump(final_output, f, indent=2, ensure_ascii=False)

    print("✅ Q&A data extracted to account_qas.json")

# Run it
process_excel_to_json("NUST Bank-Product-Knowledge.xlsx", True) 