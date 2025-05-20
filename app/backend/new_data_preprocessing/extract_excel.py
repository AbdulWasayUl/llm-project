import openpyxl
from backend.new_data_preprocessing.utils import is_question, clean_answer_list

def extract_qa_from_sheet(sheet):
    qas = []
    max_row = sheet.max_row
    max_col = sheet.max_column
    title = sheet.cell(row=1, column=1).value
    current_question = None
    current_answer = []
    row_idx = 2

    while row_idx <= max_row:
        row = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
        found_question = False

        for col_idx, cell_value in enumerate(row):
            if is_question(cell_value):
                found_question = True
                if current_question and current_answer:
                    qas.append({
                        "question": current_question.strip(),
                        "answer": clean_answer_list(current_answer)
                    })
                current_question = cell_value
                current_answer = [val for val in row[col_idx + 1:] if val is not None]
                break

        if not found_question and current_question:
            current_answer.extend([val for val in row if val is not None])
        row_idx += 1

    if current_question and current_answer:
        question_text = current_question.strip()
        if title:
            question_text += f" for {title.strip()}"
        qas.append({
            "question": question_text,
            "answer": " ".join(clean_answer_list(current_answer))
        })

    return {
        "title": title,
        "qas": qas
    }

def extract_from_excel(filepath, skip_sheets=2):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheetnames = wb.sheetnames

    all_qas = []
    for sheet_name in sheetnames[skip_sheets:]:
        sheet = wb[sheet_name]
        result = extract_qa_from_sheet(sheet)
        all_qas.extend(result["qas"])

    return {
        "questions": all_qas
    }
